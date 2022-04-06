from bs4 import BeautifulSoup
import urllib.request
from urllib.parse import urlparse
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment


class Data:
    def __init__(self):
        self.cpu = dict()
        self.ram = dict()
        self.gpu = dict()
        self.monitor = dict()
        self.price = float()
        self.points = float()


def set_data(data):
    wb = load_workbook("unsorted.xlsx")
    ws = wb.active
    last_column = 3
    for key, value in data.items():
        ws.cell(row=1, column=last_column).value = key
        # CPU
        ws.cell(row=3, column=last_column).value = value.cpu['model']
        ws.cell(row=4, column=last_column).value = value.cpu['cpus']
        ws.cell(row=5, column=last_column).value = value.cpu['frequency']
        # RAM
        ws.cell(row=7, column=last_column).value = value.ram['size']
        ws.cell(row=8, column=last_column).value = value.ram['type']
        # GPU
        ws.cell(row=10, column=last_column).value = value.gpu['model']
        ws.cell(row=11, column=last_column).value = value.gpu['type']
        # Monitor
        ws.cell(row=13, column=last_column).value = value.monitor['size']
        ws.cell(row=14, column=last_column).value = value.monitor['resolution']
        ws.cell(row=15, column=last_column).value = value.monitor['sensory']
        # Price
        ws.cell(row=17, column=last_column).value = value.price
        # Points
        ws.cell(row=18, column=last_column).value = value.points

        last_column += 1
    wb.save("unsorted.xlsx")


def prepare_sheet():
    wb = Workbook()
    ws = wb.active
    alignment_ = Alignment(horizontal='center',
                           vertical='center',
                           text_rotation=0,
                           wrap_text=False,
                           shrink_to_fit=False,
                           indent=0)

    ws['A1'] = "link"
    ws['A1'].alignment = alignment_
    ws.merge_cells('A1:B1')

    # CPU
    ws['A3'] = "CPU"
    ws['A3'].alignment = alignment_
    ws['B3'] = "Model"
    ws['B3'].alignment = alignment_
    ws['B4'] = "CPUs"
    ws['B4'].alignment = alignment_
    ws['B5'] = "Frequency"
    ws['B5'].alignment = alignment_
    ws.merge_cells('A3:A5')

    # RAM
    ws['A7'] = "RAM"
    ws['A7'].alignment = alignment_
    ws['B7'] = "Size"
    ws['B7'].alignment = alignment_
    ws['B8'] = "Type"
    ws['B8'].alignment = alignment_
    ws.merge_cells('A7:A8')

    # GPU
    ws['A10'] = "GPU"
    ws['A10'].alignment = alignment_
    ws['B10'] = "Model"
    ws['B10'].alignment = alignment_
    ws['B11'] = "Type"
    ws['B11'].alignment = alignment_
    ws.merge_cells('A10:A11')

    # Monitor
    ws['A13'] = "Monitor"
    ws['A13'].alignment = alignment_
    ws['B13'] = "Size"
    ws['B13'].alignment = alignment_
    ws['B14'] = "Resolution"
    ws['B14'].alignment = alignment_
    ws['B15'] = "Sensory?"
    ws['B15'].alignment = alignment_
    ws.merge_cells('A13:A15')

    ws['A17'] = "Price"
    ws['A17'].alignment = alignment_
    ws.merge_cells('A17:B17')

    ws['A18'] = "Points"
    ws['A18'].alignment = alignment_
    ws.merge_cells('A18:B18')

    wb.save("unsorted.xlsx")


def fitness_function(data: Data):
    c1 = 19
    c2 = 30
    c3 = 10
    c4 = 1
    c5 = 2 * 10 ** -5
    c6 = 5 * 10 ** 6
    splitted_resolution = data.monitor['resolution'].split('x')
    return c1 * data.cpu['cpus'] + c2 * data.cpu['frequency'] + c3 * data.ram['size'] + c4 * data.monitor[
        'size'] + c5 * int(splitted_resolution[0]) * int(splitted_resolution[1]) + c6 / data.price


def get_product_page(url: str):
    url += "/features"
    html_page = urllib.request.urlopen(url)
    if html_page.getcode() != 200:
        print(url, " get request failed with code ", html_page.getcode())
        return False
    soup = BeautifulSoup(html_page, "html.parser")
    data = Data()
    price = soup.find("span", {"class": "k1w"})
    if price is None:
        price = soup.find("span", {"class": "k1w wk1"})
    if price is None:
        data.price = -1.0
    else:
        price = price.text.replace(u'\u2009', '')
        price = price.replace(' ', '')
        data.price = float(price[:-2])
    features = soup.find("div", id="section-characteristics").text
    features = features.replace(" ", "")
    splitted = re.sub(r'([А-Я])', r' \1', features).split()

    for i in range(0, len(splitted)):
        if splitted[i].__contains__("Модельпроцессора"):
            data.cpu['model'] = splitted[i][16:]
        elif splitted[i].__contains__("Числоядерпроцессора"):
            data.cpu['cpus'] = int(splitted[i][19:])
        elif splitted[i].__contains__("Гц") and re.fullmatch(r'(Гц\d.\d)', splitted[i]):
            data.cpu['frequency'] = float(splitted[i][2:])
        elif splitted[i] == "Оперативнаяпамять":
            data.ram['size'] = int(splitted[i + 1][17:])
            i += 2
        elif splitted[i].__contains__("Типпамяти"):
            data.ram['type'] = splitted[i][9:]
        elif splitted[i].__contains__("Видеокарта"):
            data.gpu['model'] = splitted[i][10:]
        elif splitted[i] == "Типвидеокарты":
            data.gpu['type'] = splitted[i + 1]
        elif splitted[i].__contains__("Диагональэкрана,дюймы"):
            data.monitor['size'] = float(splitted[i][21:])
        elif splitted[i].__contains__("Технологияматрицы"):
            data.monitor['type'] = splitted[i][17:]
        elif splitted[i].__contains__("Разрешениеэкрана"):
            data.monitor['resolution'] = splitted[i][16:]
        elif splitted[i] == "Сенсорныйэкран":
            if splitted[i + 1] == "Нет":
                data.monitor['sensory'] = False
            else:
                data.monitor['sensory'] = True
    data.points = fitness_function(data)
    print("Fitness:", data.points)
    return data


allData = dict()


def get_list_page(url: str):
    global allData
    parsed_link = urlparse(url)
    html_page = urllib.request.urlopen(url)
    if html_page.getcode() != 200:
        print(url, " get request failed with code ", html_page.getcode())
        return False
    soup = BeautifulSoup(html_page, "html.parser")
    for link_iter in soup.findAll('a'):
        if str(link_iter.get('href')).__contains__('/product'):
            print(parsed_link.hostname + link_iter.get('href'))
            if not allData.__contains__(parsed_link.hostname + link_iter.get('href')):
                allData[parsed_link.hostname + link_iter.get('href')] = get_product_page(
                    "https://" + parsed_link.hostname + link_iter.get('href'))
                if len(allData) == 1:
                    set_data(allData)
                    return False
    return True


prepare_sheet()
link = "https://www.ozon.ru/category/noutbuki-15692/"
second_page = True
page_number = 2
while True:
    if not get_list_page(link):
        break
    if second_page:
        link += "page=" + str(page_number)
        second_page = False
    else:
        link = link[:len(link) - 2] + str(page_number)
    page_number += 1
